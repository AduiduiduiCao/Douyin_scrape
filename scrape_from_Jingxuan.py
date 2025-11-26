import os
import re
import json
import time
import random
from urllib.parse import unquote, urlparse
from typing import Any, Dict, Optional, List

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, WebDriverException

# ====== 配置 ======
EXCEL_PATH = "/Users/punic/douyin_video_stats/target_douyinURL.xlsx"
LINK_COL_LETTER = "M"  # 发布链接列
JSON_OUTPUT = "/Users/punic/douyin_video_stats/Douyin_excel_analysis.json"
DEBUG_RENDERDATA_DIR = "/Users/punic/douyin_video_stats/renderdata_debug"

LIKE_COL = "O"
COMMENT_COL = "P"
SHARE_COL = "Q"
COLLECT_COL = "R"

SLEEP_MIN = 3.0
SLEEP_MAX = 7.0


# ====== 工具函数 ======
def ensure_dir(path: str):
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def parse_count_text(text: str) -> int:
    """
    把页面上显示的“点赞数”等文案转成整数：
    - "1234" -> 1234
    - "1.2万" / "1.2w" / "1.2W" -> 12000
    """
    if not text:
        return 0
    s = text.strip()

    # 去掉多余空格和逗号
    s = s.replace(",", "").replace("，", "")

    # 特殊情况
    if s in {"点赞", "评论", "分享", "收藏", "-", "—"}:
        return 0

    # 万制
    if s.endswith("万") or s.lower().endswith("w"):
        num_part = s[:-1]
        try:
            val = float(num_part)
            return int(val * 10000)
        except ValueError:
            return 0

    # 纯数字
    try:
        return int(s)
    except ValueError:
        # 万一有奇怪格式，就尽量提取数字
        m = re.search(r"\d+(\.\d+)?", s)
        if not m:
            return 0
        num_txt = m.group(0)
        try:
            if "." in num_txt:
                return int(float(num_txt))
            return int(num_txt)
        except ValueError:
            return 0


# ====== 浏览器相关 ======
def init_driver() -> webdriver.Chrome:
    print("[*] 正在启动 Chrome…")
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()

    driver.get("https://www.douyin.com")
    print("[-] Chrome 已打开抖音首页，请在 Chrome 中手动完成 登录/扫码/短信验证。")
    input(">>> 登录完成后，在此终端按 Enter 继续... \n")
    return driver


# ====== Excel 相关 ======
def ensure_stat_columns(ws):
    header_row = 1

    def ensure_column(col_letter: str, title: str):
        cell = ws[f"{col_letter}{header_row}"]
        if cell.value is None:
            cell.value = title
            print(f"[*] 在列 {col_letter} 新增列标题: {title}")
        else:
            print(f"[*] 已存在列 {col_letter}：{cell.value}")

    ensure_column(LIKE_COL, "点赞")
    ensure_column(COMMENT_COL, "评论")
    ensure_column(SHARE_COL, "分享")
    ensure_column(COLLECT_COL, "收藏")


def extract_first_url(text: str) -> Optional[str]:
    if not text:
        return None
    text = str(text).strip()
    m = re.search(r"https?://\S+", text)
    if not m:
        return None
    url = m.group(0)
    url = url.rstrip("，。！？!?,，）)」》>」」")
    return url


# ====== RENDER_DATA 解析 ======
def get_render_data_json(driver: webdriver.Chrome, row_idx: int) -> Optional[Dict[str, Any]]:
    """从 meta 或 script#RENDER_DATA 里解析 JSON。"""
    # 先 meta
    try:
        meta_el = driver.find_element(By.CSS_SELECTOR, 'meta[name="RENDER_DATA"]')
        content = meta_el.get_attribute("content") or ""
    except NoSuchElementException:
        content = ""

    if content:
        s = content.strip()
        if s.startswith("%7B"):
            s = unquote(s)
        try:
            data = json.loads(s)
            print(f"    [行 {row_idx}] 通过 meta[RENDER_DATA] 解析成功。")
            return data
        except Exception as e:
            print(f"    [行 {row_idx}] meta[RENDER_DATA] JSON 解析失败: {e}")
            print(f"        前 120 字符预览: {s[:120]!r}")

    # 再 script#RENDER_DATA
    try:
        script_el = driver.find_element(By.CSS_SELECTOR, "script#RENDER_DATA")
        s = (
            script_el.get_attribute("innerHTML")
            or script_el.get_attribute("textContent")
            or ""
        ).strip()
    except NoSuchElementException:
        s = ""

    if not s:
        print(f"    [行 {row_idx}] 未找到 meta[RENDER_DATA] 或 script#RENDER_DATA。")
        return None

    if s.startswith("%7B"):
        s = unquote(s)

    try:
        data = json.loads(s)
        print(f"    [行 {row_idx}] 通过 script#RENDER_DATA 解析成功。")
        if isinstance(data, dict):
            print(f"        顶层 key 预览: {list(data.keys())[:5]}")
        return data
    except Exception as e:
        print(f"    [行 {row_idx}] script#RENDER_DATA JSON 解析失败: {e}")
        print(f"        前 120 字符预览: {s[:120]!r}")
        return None


def debug_dump_renderdata(data: Dict[str, Any], aweme_id: str, row_idx: int):
    """把整棵 RENDER_DATA dump 到本地文件，方便后续手动分析。"""
    try:
        ensure_dir(DEBUG_RENDERDATA_DIR)
        fname = os.path.join(
            DEBUG_RENDERDATA_DIR, f"renderdata_row{row_idx}_{aweme_id}.json"
        )
        with open(fname, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"    [行 {row_idx}] 已将 RENDER_DATA dump 到: {fname}")
    except Exception as e:
        print(f"    [行 {row_idx}] dump RENDER_DATA 失败: {e}")


def find_stats_in_json(obj: Any, depth: int = 0) -> Optional[Dict[str, int]]:
    """旧方案：递归 JSON，找包含 diggCount/commentCount/shareCount/collectCount 的对象。"""
    if depth > 30:
        return None

    if isinstance(obj, dict):
        keys = set(obj.keys())
        wanted = {"diggCount", "commentCount", "shareCount", "collectCount"}
        if wanted.issubset(keys):
            try:
                return {
                    "digg": int(obj.get("diggCount", 0) or 0),
                    "comment": int(obj.get("commentCount", 0) or 0),
                    "share": int(obj.get("shareCount", 0) or 0),
                    "collect": int(obj.get("collectCount", 0) or 0),
                }
            except Exception:
                pass

        # 优先深入 stats/statistics
        for special_key in ("stats", "statistics"):
            if special_key in obj and isinstance(obj[special_key], (dict, list)):
                found = find_stats_in_json(obj[special_key], depth + 1)
                if found:
                    return found

        for v in obj.values():
            found = find_stats_in_json(v, depth + 1)
            if found:
                return found

    elif isinstance(obj, list):
        for item in obj:
            found = find_stats_in_json(item, depth + 1)
            if found:
                return found

    return None


# ====== DOM 抓统计（新方案，优先使用） ======
def try_scrape_stats_from_dom(driver: webdriver.Chrome, row_idx: int) -> Optional[Dict[str, int]]:
    """
    尝试直接从页面上抓点赞/评论/分享/收藏：
    依赖 data-e2e 属性（Douyin PC 上通常存在），
    如果任一指标找不到就视为失败。
    """
    # 不同指标尝试的 selector 列表（都试一轮，先匹配到的为准）
    selectors = {
        "digg": [
            '[data-e2e="like-count"]',
            '[data-e2e="like-icon"] + span',
        ],
        "comment": [
            '[data-e2e="comment-count"]',
        ],
        "share": [
            '[data-e2e="share-count"]',
        ],
        "collect": [
            '[data-e2e="favorite-count"]',
            '[data-e2e="collect-count"]',
        ],
    }

    result: Dict[str, int] = {}
    all_ok = True

    for key, css_list in selectors.items():
        val: Optional[int] = None
        for css in css_list:
            try:
                el = driver.find_element(By.CSS_SELECTOR, css)
                txt = (el.text or "").strip()
                if not txt:
                    continue
                val = parse_count_text(txt)
                print(f"        [DOM] {key} via {css} -> '{txt}' -> {val}")
                break
            except NoSuchElementException:
                continue
        if val is None:
            print(f"        [DOM 调试] 未找到 {key} 对应元素（selectors={css_list}）。")
            all_ok = False
            break
        result[key] = val

    if not all_ok:
        return None

    print(
        f"    [行 {row_idx}] DOM 抓取统计：点赞 {result['digg']} 评论 {result['comment']} "
        f"分享 {result['share']} 收藏 {result['collect']}"
    )
    return result


# ====== 打开视频页 ======
def goto_video_page(driver: webdriver.Chrome, url: str, row_idx: int) -> Optional[str]:
    """
    打开 Excel 里的链接，让浏览器自己跳转。
    如果落在 iesdouyin 分享页，尝试拼 PC 视频页。
    返回最终 current_url。
    """
    try:
        driver.get(url)
    except WebDriverException as e:
        print(f"    [行 {row_idx}] 打开链接失败: {e}")
        return None

    time.sleep(2.5)
    current = driver.current_url
    print(f"    [行 {row_idx}] 首次落地 URL: {current}")

    parsed = urlparse(current)

    # iesdouyin 分享页 -> 转 PC 视频页
    if "iesdouyin.com" in parsed.netloc and "/share/video/" in parsed.path:
        parts = parsed.path.rstrip("/").split("/")
        aweme_id = parts[-1]
        if aweme_id.isdigit():
            target = f"https://www.douyin.com/video/{aweme_id}"
            print(f"    [行 {row_idx}] 从 iesdouyin 分享页跳转到 PC 视频页: {target}")
            try:
                driver.get(target)
                time.sleep(2.5)
                current = driver.current_url
            except WebDriverException as e:
                print(f"    [行 {row_idx}] 跳转 PC 视频页失败: {e}")
                return None

    return current


def extract_aweme_id_from_url(url: str) -> str:
    parsed = urlparse(url)
    parts = parsed.path.rstrip("/").split("/")
    if len(parts) >= 3 and parts[-2] == "video":
        return parts[-1]
    return "unknown"


# ====== 针对单条 URL 抓统计 ======
def fetch_stats_for_one_url(
    driver: webdriver.Chrome, url: str, row_idx: int
) -> Optional[Dict[str, int]]:
    print(f"[行 {row_idx}] 提取到链接: {url}")
    final_url = goto_video_page(driver, url, row_idx)
    if not final_url:
        return None

    # 1. 先尝试 DOM 抓取
    dom_stats = try_scrape_stats_from_dom(driver, row_idx)
    if dom_stats:
        return dom_stats

    # 2. DOM 失败，退回 RENDER_DATA 方案，同时 dump 一份 JSON 方便你后面分析结构
    data = get_render_data_json(driver, row_idx)
    if not data:
        return None

    aweme_id = extract_aweme_id_from_url(final_url)
    debug_dump_renderdata(data, aweme_id, row_idx)

    stats = find_stats_in_json(data)
    if not stats:
        print(f"    [行 {row_idx}] 在 RENDER_DATA 中未找到统计信息。")
        return None

    print(
        f"    [行 {row_idx}] RENDER_DATA 抓取统计：点赞 {stats['digg']} 评论 {stats['comment']} "
        f"分享 {stats['share']} 收藏 {stats['collect']}"
    )
    return stats


# ====== 主流程：Excel 遍历 + 写入 ======
def process_excel(driver: webdriver.Chrome):
    print(f"[*] 正在打开 Excel: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    link_header = ws[f"{LINK_COL_LETTER}1"].value
    print(f"[*] 默认使用 {LINK_COL_LETTER} 列作为抖音链接列，表头：{link_header}")
    ensure_stat_columns(ws)

    max_row = ws.max_row
    print(f"[*] 检测到总行数: {max_row}（包含表头）")

    results: List[Dict[str, Any]] = []

    for row in range(2, max_row + 1):
        cell = ws[f"{LINK_COL_LETTER}{row}"].value
        if not cell:
            print(f"[行 {row}] {LINK_COL_LETTER} 列为空，跳过")
            continue

        url = extract_first_url(str(cell))
        if not url:
            print(f"[行 {row}] 未在文本中找到有效链接，原始内容: {str(cell)[:50]}...")
            continue

        stats = fetch_stats_for_one_url(driver, url, row)
        sleep_time = random.uniform(SLEEP_MIN, SLEEP_MAX)
        print(f"    [行 {row}] 暂停 {sleep_time:.1f} 秒，以降低访问频率…")
        time.sleep(sleep_time)

        if not stats:
            continue

        ws[f"{LIKE_COL}{row}"].value = stats["digg"]
        ws[f"{COMMENT_COL}{row}"].value = stats["comment"]
        ws[f"{SHARE_COL}{row}"].value = stats["share"]
        ws[f"{COLLECT_COL}{row}"].value = stats["collect"]

        results.append(
            {
                "row": row,
                "orig_text": str(cell),
                "url": url,
                "stats": stats,
            }
        )

    # 写 JSON 方便你后续分析
    try:
        with open(JSON_OUTPUT, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"[*] 共成功解析 {len(results)} 条视频，已写入 JSON: {JSON_OUTPUT}")
    except Exception as e:
        print(f"[!] 写入 JSON 时出错: {e}")

    # 保存 Excel
    try:
        wb.save(EXCEL_PATH)
        print(f"[*] 已保存所有修改到 Excel: {EXCEL_PATH}")
    except PermissionError:
        print(f"[!] 保存 Excel 失败：没有权限写入 {EXCEL_PATH}，请检查文件是否被其他程序占用。")
    except Exception as e:
        print(f"[!] 保存 Excel 时发生错误: {e}")


def main():
    driver = None
    try:
        driver = init_driver()
        process_excel(driver)
    finally:
        if driver is not None:
            driver.quit()
            print("[*] 已关闭浏览器，程序结束。")


if __name__ == "__main__":
    main()